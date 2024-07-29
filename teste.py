import os
import pandas as pd
import openpyxl
from PyPDF2 import PdfReader
import re

def extract_codes_from_pdf(pdf_path):
    with open(pdf_path, 'rb') as f:
        reader = PdfReader(f)
        num_pages = len(reader.pages)
        codes = set()
        for page_num in range(num_pages):
            page = reader.pages[page_num]
            text = page.extract_text()
            # Encontrar todos os códigos de instalação
            found_codes = re.findall(r'Instalação:\s*(\d{10})', text)
            # Remover zeros à esquerda
            found_codes = [code.lstrip('0') for code in found_codes]
            codes.update(found_codes)
        print(codes)
        return list(codes)

def find_and_update_excel(excel_path, codes, not_found_path, pdf_name):
    wb            = openpyxl.load_workbook(excel_path)
    ws            = wb.active
    codes_found   = set()
    colors        = ["FFFF00", "FFA07A", "20B2AA", "87CEFA", "9370DB", "3CB371", "FFB6C1", "8B4513", "708090", "6A5ACD", "4682B4", "D2B48C", "008080"]
    summary_lines = []

    # Obter cores já utilizadas nas células
    used_colors = set()
    for row in ws.iter_rows():
        for cell in row:
            if cell.fill.start_color.index != '00000000' and cell.fill.start_color.index != '000000':  # Ignorar células sem cor
                used_colors.add(cell.fill.start_color.rgb)

    color_index = 0
    color = colors[color_index % len(colors)]


    for code in codes:
        found = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == code:
                    if cell.fill.start_color.rgb == '00000000' or cell.fill.start_color.rgb == '000000':
                        while color in used_colors:
                            color_index += 1
                            color = colors[color_index % len(colors)]
                        cell.fill = openpyxl.styles.PatternFill(start_color=color, end_color=color, fill_type="solid")
                        codes_found.add(cell.value)
                        used_colors.add(color)
                        print(codes_found)
                        found = True

        if found:
            continue


    codes_not_found = set(codes) - codes_found

    summary_lines.append(f"Fatura: {pdf_name}")
    summary_lines.append(f"Quantidade de faturas: {len(codes)}")
    summary_lines.append(f"Cor utilizada: {color}")
    summary_lines.append(f"Códigos ausentes: {', '.join(codes_not_found) if codes_not_found else 'Nenhum'}")
    summary_lines.append(f"Faturas não encontradas: {len(codes_not_found)}")
    summary_lines.append("")

    with open("teste.txt", 'w') as f:
        f.write("\n".join(summary_lines))

    wb.save(excel_path)


def main(pdf_path, excel_path):
    if not os.path.isfile(pdf_path):
        print(f"Erro: Arquivo PDF '{pdf_path}' não encontrado.")
        return

    if not os.path.isfile(excel_path):
        print(f"Erro: Arquivo Excel '{excel_path}' não encontrado.")
        return

    pdf_name = pdf_path.split("/")[-1]
    not_found_path = pdf_path.replace('.pdf', '.txt')
    codes = extract_codes_from_pdf(pdf_path)
    if codes:
        find_and_update_excel(excel_path, codes, not_found_path, pdf_name)
        print("Processo concluído com sucesso!")
    else:
        print(f"Não foram encontrados códigos de instalação no PDF {pdf_name}")

if __name__ == "__main__":
    pdf_path = ".venv/Conta - 400000010110.PDF"
    excel_path = ".venv/excelteste.xlsx"

    main(pdf_path, excel_path)
