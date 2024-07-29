import os
import pandas as pd
import openpyxl
from PyPDF2 import PdfReader
import re
import tkinter as tk
from tkinter import filedialog, messagebox

def extract_codes_from_pdf(pdf_path):
    try:
        with open(pdf_path, 'rb') as f:
            reader = PdfReader(f)
            num_pages = len(reader.pages)
            codes = set()
            for page_num in range(num_pages):
                page = reader.pages[page_num]
                text = page.extract_text()
                # Encontrar todos os códigos de instalação
                found_codes = re.findall(r'Instalação:\s*(\d{10})', text)
                # Remover zeros à esquerda
                found_codes = [code.lstrip('0') for code in found_codes]
                codes.update(found_codes)
            return list(codes)
    except Exception as e:
        messagebox.showerror("Erro ao extrair códigos", f"Erro ao processar o arquivo PDF '{pdf_path}': {str(e)}")
        return []

def find_and_update_excel(excel_path, pdf_details, summary_path):
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        colors = ["FFFF00", "FFA07A", "20B2AA", "87CEFA", "9370DB", "3CB371", "FFB6C1", "8B4513", "708090", "6A5ACD",
                  "4682B4", "D2B48C", "008080"]
        summary_lines = []

        for index, (pdf_name, codes) in enumerate(pdf_details.items()):
            color = colors[index % len(colors)]
            codes_found = set()

            for code in codes:
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value == code:
                            cell.fill = openpyxl.styles.PatternFill(start_color=color, end_color=color, fill_type="solid")
                            codes_found.add(cell.value)
                            break

            codes_not_found = set(codes) - codes_found
            summary_lines.append(f"Fatura: {pdf_name}")
            summary_lines.append(f"Quantidade de faturas: {len(codes)}")
            summary_lines.append(f"Cor utilizada: {color}")
            summary_lines.append(f"Códigos ausentes: {', '.join(codes_not_found) if codes_not_found else 'Nenhum'}")
            summary_lines.append(f"Faturas não encontradas: {len(codes_not_found)}")
            summary_lines.append("")

        with open(summary_path, 'w') as f:
            f.write("\n".join(summary_lines))

        wb.save(excel_path)
    except Exception as e:
        messagebox.showerror("Erro ao atualizar Excel", f"Erro ao atualizar o arquivo Excel '{excel_path}': {str(e)}")

def main():
    # Criar a janela principal
    root = tk.Tk()
    root.withdraw()  # Esconder a janela principal

    # Selecionar arquivos PDF
    pdf_paths = filedialog.askopenfilenames(title="Selecione os arquivos PDF", filetypes=[("PDF files", "*.pdf")])
    if not pdf_paths:
        messagebox.showerror("Erro", "Nenhum arquivo PDF selecionado.")
        return

    # Selecionar arquivo Excel
    excel_path = filedialog.askopenfilename(title="Selecione o arquivo Excel", filetypes=[("Excel files", "*.xlsx")])
    if not excel_path:
        messagebox.showerror("Erro", "Nenhum arquivo Excel selecionado.")
        return

    if not os.path.isfile(excel_path):
        messagebox.showerror("Erro", f"Arquivo Excel '{excel_path}' não encontrado.")
        return

    pdf_details = {}
    for pdf_path in pdf_paths:
        if not os.path.isfile(pdf_path):
            messagebox.showerror("Erro", f"Arquivo PDF '{pdf_path}' não encontrado.")
            continue
        pdf_name = os.path.basename(pdf_path)
        codes = extract_codes_from_pdf(pdf_path)
        if codes:
            pdf_details[pdf_name] = codes
        else:
            messagebox.showwarning("Aviso", f"Não foram encontrados códigos de instalação no PDF {pdf_name}")

    if pdf_details:
        summary_path = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Text files", "*.txt")],
                                                    title="Salvar relatório como")
        if not summary_path:
            messagebox.showerror("Erro", "Nenhum arquivo de relatório selecionado.")
            return
        find_and_update_excel(excel_path, pdf_details, summary_path)
        messagebox.showinfo("Sucesso", "Processo concluído com sucesso!")
    else:
        messagebox.showwarning("Aviso", "Nenhum código de instalação foi encontrado em todas as faturas fornecidas.")

if __name__ == "__main__":
    main()
